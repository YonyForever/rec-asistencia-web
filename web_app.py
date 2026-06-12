import os
import sys
import datetime
import base64
import io
import threading
import numpy as np
import cv2
from PIL import Image
import pyotp
import telebot
from fastapi import FastAPI, HTTPException
from fastapi.responses import HTMLResponse, FileResponse
from pydantic import BaseModel
from contextlib import asynccontextmanager  # <--- IMPORTANTE

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_DISPONIBLE = True
except ImportError:
    EXCEL_DISPONIBLE = False

from core.config import (
    TOTP_INTERVAL,
    TELEGRAM_BOT_TOKEN,
    TELEGRAM_CHAT_ID,
    BASE_DIR
)
from core.face_engine import FaceEngine
from core.qr_engine import QREngine
from core.model_loader import ensure_models_exist
from database.connection import get_session, init_db
from database.models import Alumno, ClaseConfig, Sesion, Asistencia

# Variables globales de los motores (se instanciarán al arrancar la app)
face_engine = None
qr_engine = None

def seed_clases():
    """Inserta materias de ejemplo si la tabla está vacía (primer despliegue)."""
    session = get_session()
    try:
        count = session.query(ClaseConfig).count()
        if count == 0:
            clases_ejemplo = [
                ClaseConfig(nombre_materia="Gestión de Proyectos", hora_inicio="08:00", limite_presente=5, limite_tarde=20, activo=1),
                ClaseConfig(nombre_materia="Base de Datos", hora_inicio="10:00", limite_presente=5, limite_tarde=20, activo=1),
                ClaseConfig(nombre_materia="Programación Web", hora_inicio="12:00", limite_presente=5, limite_tarde=20, activo=1),
                ClaseConfig(nombre_materia="Redes y Comunicaciones", hora_inicio="14:00", limite_presente=5, limite_tarde=20, activo=1),
                ClaseConfig(nombre_materia="Inteligencia Artificial", hora_inicio="16:00", limite_presente=5, limite_tarde=20, activo=1),
            ]
            session.add_all(clases_ejemplo)
            session.commit()
            print("[Seed] 5 materias de ejemplo insertadas en la BD.")
        else:
            print(f"[Seed] Ya existen {count} materias. No se insertaron datos semilla.")
    except Exception as e:
        session.rollback()
        print(f"[Seed ERROR] {e}")
    finally:
        session.close()


# Definir el ciclo de vida de la aplicación para entornos Cloud (Render)
@asynccontextmanager
async def lifespan(app: FastAPI):
    global face_engine, qr_engine
    print("[Lifespan] Iniciando secuencia de arranque en el servidor...")
    
    # 1. Inicializar la Base de Datos (crear tablas)
    init_db()
    
    # 2. Sembrar datos iniciales si la BD está vacía
    seed_clases()
    
    # 3. Descargar modelos de IA de forma segura (Render no cortará la conexión aquí)
    ensure_models_exist()
    
    # 4. Instanciar motores una vez descargados los archivos ONNX
    face_engine = FaceEngine()
    face_engine.reload_students_cache()
    qr_engine = QREngine()
    
    print("[Lifespan] Motores biométricos listos para operar.")
    yield
    print("[Lifespan] Apagando servidor...")

# Pasar el lifespan a la instancia de FastAPI
app = FastAPI(title="REC — API de Asistencia Biométrica", lifespan=lifespan)

# Estado Global de la Sesión Activa
active_session_id = None
active_class_name = ""
hora_apertura = None
local_limite_presente = 5
local_limite_tarde = 20
session_active = False

class StartSessionRequest(BaseModel):
    clase_config_id: int
    limite_presente: int = 5
    limite_tarde: int = 20

class ScanRequest(BaseModel):
    image: str  # Imagen codificada en Base64 (data:image/jpeg;base64,...)

class EnrollScanRequest(BaseModel):
    image: str

class EnrollSaveRequest(BaseModel):
    alumno_id: str
    nombre: str
    image: str

class CreateClassRequest(BaseModel):
    nombre_materia: str
    hora_inicio: str
    limite_presente: int = 5
    limite_tarde: int = 20


@app.get("/", response_class=HTMLResponse)
def get_index():
    """Sirve la interfaz web principal del docente."""
    return FileResponse("templates/index.html")


@app.get("/enroll", response_class=HTMLResponse)
def get_enroll_page():
    """Sirve la interfaz de enrolamiento de alumnos."""
    return FileResponse("templates/enroll.html")


@app.get("/student", response_class=HTMLResponse)
def get_student_page():
    """Sirve la app móvil/web del alumno para generar el QR dinámico."""
    return FileResponse("templates/student.html")


@app.post("/api/enroll/scan")
def api_enroll_scan(req: EnrollScanRequest):
    """Detecta si hay un rostro en escena durante el enrolamiento."""
    try:
        header, encoded = req.image.split(",", 1)
        image_data = base64.b64decode(encoded)
        image = Image.open(io.BytesIO(image_data))
        frame = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)

        h, w, _ = frame.shape
        face_engine.detector.setInputSize((w, h))
        _, faces = face_engine.detector.detect(frame)

        if faces is not None and len(faces) > 0:
            box = list(map(int, faces[0][0:4]))
            return {"detected": True, "box": box}
        return {"detected": False}
    except Exception as e:
        return {"detected": False, "error": str(e)}


@app.post("/api/enroll/save")
def api_enroll_save(req: EnrollSaveRequest):
    """Genera la firma biométrica del rostro y registra al alumno en la BD."""
    alumno_id = req.alumno_id.strip()
    nombre = req.nombre.strip()
    if not alumno_id or not nombre:
        return {"status": "error", "message": "ID y Nombre son obligatorios."}

    session = get_session()
    try:
        existente = session.query(Alumno).filter_by(id=alumno_id).first()
        if existente:
            return {"status": "error", "message": f"El ID {alumno_id} ya se encuentra registrado."}

        header, encoded = req.image.split(",", 1)
        image_data = base64.b64decode(encoded)
        image = Image.open(io.BytesIO(image_data))
        frame = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)

        h, w, _ = frame.shape
        face_engine.detector.setInputSize((w, h))
        _, faces = face_engine.detector.detect(frame)

        if faces is None or len(faces) == 0:
            return {"status": "error", "message": "No se detecta rostro en el encuadre."}

        embedding = face_engine.extract_embedding(frame, faces[0])

        # Evitar registrar caras que ya pertenecen a otros estudiantes
        m_id, m_name, _ = face_engine.match_face(embedding)
        if m_id:
            return {"status": "error", "message": f"Este rostro ya pertenece al alumno registrado '{m_name}' ({m_id})."}

        totp_secret = pyotp.random_base32()

        nuevo_alumno = Alumno(id=alumno_id, nombre=nombre)
        nuevo_alumno.set_embedding(embedding)
        nuevo_alumno.totp_secret = totp_secret
        session.add(nuevo_alumno)
        session.commit()

        # Recargar caché en memoria
        face_engine.reload_students_cache()

        totp = pyotp.TOTP(totp_secret, interval=TOTP_INTERVAL)
        uri = totp.provisioning_uri(name=alumno_id, issuer_name="REC-Asistencia")

        return {
            "status": "ok",
            "totp_secret": totp_secret,
            "uri": uri
        }

    except Exception as e:
        session.rollback()
        return {"status": "error", "message": str(e)}
    finally:
        session.close()



@app.get("/api/clases")
def get_clases():
    """Devuelve las clases/materias configuradas en la BD."""
    session = get_session()
    try:
        clases = session.query(ClaseConfig).filter_by(activo=1).all()
        return [
            {
                "id": c.id,
                "nombre_materia": c.nombre_materia,
                "hora_inicio": c.hora_inicio
            }
            for c in clases
        ]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        session.close()


@app.post("/api/clases")
def create_clase(req: CreateClassRequest):
    """Crea una nueva materia/clase en la BD."""
    session = get_session()
    try:
        nueva = ClaseConfig(
            nombre_materia=req.nombre_materia.strip(),
            hora_inicio=req.hora_inicio.strip(),
            limite_presente=req.limite_presente,
            limite_tarde=req.limite_tarde,
            activo=1
        )
        session.add(nueva)
        session.commit()
        return {"status": "ok", "id": nueva.id, "nombre_materia": nueva.nombre_materia}
    except Exception as e:
        session.rollback()
        return {"status": "error", "message": str(e)}
    finally:
        session.close()


@app.delete("/api/clases/{clase_id}")
def delete_clase(clase_id: int):
    """Desactiva (soft delete) una materia."""
    session = get_session()
    try:
        clase = session.query(ClaseConfig).filter_by(id=clase_id).first()
        if not clase:
            return {"status": "error", "message": "Clase no encontrada."}
        clase.activo = 0
        session.commit()
        return {"status": "ok"}
    except Exception as e:
        session.rollback()
        return {"status": "error", "message": str(e)}
    finally:
        session.close()


@app.post("/api/session/start")
def start_session(req: StartSessionRequest):
    global active_session_id, active_class_name, hora_apertura, session_active
    global local_limite_presente, local_limite_tarde

    # 🔄 ¡AUTOREPARACIÓN! Si ya hay una sesión activa, la cerramos automáticamente antes de abrir la nueva
    if session_active:
        print("[SISTEMA] Detectada sesión antigua colgada. Forzando cierre automático...")
        try:
            # Reutiliza tu lógica existente de guardado o simplemente resetea el estado
            session = get_session()
            sesion_antigua = session.query(Sesion).filter_by(id=active_session_id).first()
            if sesion_antigua and sesion_antigua.hora_fin is None:
                sesion_antigua.hora_fin = datetime.datetime.now().time().strftime("%H:%M:%S")
                session.commit()
            session.close()
        except Exception as e:
            print(f"[ALERTA] No se pudo cerrar la sesión anterior limpiamente en BD: {e}")
        
        # Reseteamos las variables para limpiar la memoria RAM
        session_active = False

    # ── Ahora el flujo continúa normalmente sin trabarse ───────────────────
    session = get_session()
    config = session.query(ClaseConfig).filter_by(id=req.clase_config_id, activo=1).first()
    if not config:
        session.close()
        raise HTTPException(status_code=404, detail="Configuración de clase no encontrada o inactiva.")

    now = datetime.datetime.now()
    nueva_sesion = Sesion(
        clase_config_id=config.id,
        fecha=now.date(),
        hora_inicio=now.time().strftime("%H:%M:%S")
    )
    session.add(nueva_sesion)
    session.commit()

    active_session_id = nueva_sesion.id
    active_class_name = config.nombre_materia
    hora_apertura = now
    local_limite_presente = req.limite_presente
    local_limite_tarde = req.limite_tarde
    session_active = True

    # Cargar alumnos a la caché del motor de reconocimiento
    face_engine.reload_students_cache()

    session.close()
    return {
        "status": "success",
        "message": f"Sesión iniciada para {config.nombre_materia}",
        "session_id": active_session_id
    }

@app.post("/api/session/stop")
def stop_session():
    """Finaliza la sesión actual de asistencia y genera reportes."""
    global active_session_id, session_active
    
    if not session_active:
        return {"status": "error", "message": "No hay ninguna sesión activa para cerrar."}

    session = get_session()
    try:
        sesion_db = session.query(Sesion).filter_by(id=active_session_id).first()
        if sesion_db:
            sesion_db.hora_cierre = datetime.datetime.now()
            sesion_db.estado = "CERRADA"
            session.commit()

        # Generar reportes
        generate_and_send_report_web()

    except Exception as e:
        session.rollback()
        return {"status": "error", "message": str(e)}
    finally:
        session.close()

    session_active = False
    active_session_id = None
    return {"status": "ok"}


@app.post("/api/scan")
def scan_frame(req: ScanRequest):
    """Recibe un frame en Base64, realiza detección y guarda asistencia si hay match."""
    global active_session_id, session_active, hora_apertura, local_limite_presente, local_limite_tarde
    
    if not session_active or not active_session_id:
        return {"detected": False, "message": "Sesión inactiva."}

    try:
        # Decodificar imagen Base64
        header, encoded = req.image.split(",", 1)
        image_data = base64.b64decode(encoded)
        image = Image.open(io.BytesIO(image_data))
        frame = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)

        alumno_id = None
        nombre = None
        metodo = None
        exito = False
        mensaje = ""
        score = 0.0

        # 1. Probar Detección de Código QR
        decoded_info, points = qr_engine.detect_and_decode(frame)
        if decoded_info:
            success, q_id, q_name, msg = qr_engine.validate_student_qr(decoded_info)
            metodo = "QR"
            alumno_id = q_id
            nombre = q_name
            exito = success
            mensaje = msg
        else:
            # 2. Probar Detección Facial
            face_data = face_engine.detect_face(frame)
            if face_data is not None:
                embedding = face_engine.extract_embedding(frame, face_data)
                f_id, f_name, f_score = face_engine.match_face(embedding)
                metodo = "FACIAL"
                score = f_score
                if f_id:
                    alumno_id = f_id
                    nombre = f_name
                    exito = True
                else:
                    exito = False

        if not exito:
            return {
                "detected": True,
                "exito": False,
                "metodo": metodo,
                "score": score,
                "mensaje": mensaje
            }

        # Registrar Asistencia en la Base de Datos si hace Match
        now = datetime.datetime.now()
        elapsed = (now - hora_apertura).total_seconds() / 60.0

        if elapsed <= local_limite_presente:
            estado = "PRESENTE"
        elif elapsed <= local_limite_tarde:
            estado = "RETARDO"
        else:
            estado = "FALTA"

        session = get_session()
        try:
            existente = session.query(Asistencia).filter_by(
                sesion_id=active_session_id,
                alumno_id=alumno_id
            ).first()

            if not existente:
                nueva_asis = Asistencia(
                    sesion_id=active_session_id,
                    alumno_id=alumno_id,
                    metodo=metodo,
                    estado=estado
                )
                session.add(nueva_asis)
                session.commit()

            # Obtener datos de actualización para el frontend
            metrics = get_current_metrics_db(session, active_session_id)
            attendance_list = get_attendance_list_db(session, active_session_id)

            return {
                "detected": True,
                "exito": True,
                "alumno_id": alumno_id,
                "nombre": nombre,
                "metodo": metodo,
                "estado": estado,
                "metrics": metrics,
                "attendance_list": attendance_list
            }

        except Exception as db_err:
            session.rollback()
            return {"detected": False, "message": f"Error BD: {db_err}"}
        finally:
            session.close()

    except Exception as e:
        return {"detected": False, "message": f"Error de escaneo: {e}"}


def get_current_metrics_db(session, sesion_id):
    """Obtiene conteos de asistencia para la sesión actual."""
    total_alumnos = session.query(Alumno).count()
    asistentes = session.query(Asistencia).filter_by(sesion_id=sesion_id).all()
    
    presentes = sum(1 for a in asistentes if a.estado == "PRESENTE")
    retardos = sum(1 for a in asistentes if a.estado in ("TARDE", "RETARDO"))
    ausentes = total_alumnos - (presentes + retardos)
    
    return {
        "presentes": presentes,
        "retardos": retardos,
        "ausentes": max(0, ausentes)
    }


def get_attendance_list_db(session, sesion_id):
    """Obtiene la lista completa de asistencia registrada para la sesión."""
    registros = (
        session.query(Asistencia, Alumno)
        .join(Alumno, Asistencia.alumno_id == Alumno.id)
        .filter(Asistencia.sesion_id == sesion_id)
        .order_by(Asistencia.timestamp.desc())
        .all()
    )
    return [
        {
            "alumno_id": alum.id,
            "nombre": alum.nombre,
            "hora": asis.timestamp.strftime("%H:%M:%S"),
            "metodo": asis.metodo,
            "estado": asis.estado
        }
        for asis, alum in registros
    ]


def generate_and_send_report_web():
    """Registra ausencias y envía los reportes automáticos."""
    global active_session_id, active_class_name, hora_apertura
    
    session = get_session()
    try:
        sesion_db = session.query(Sesion).filter_by(id=active_session_id).first()
        if not sesion_db:
            return

        # 1. Registrar ausentes automáticamente
        todos_alumnos = session.query(Alumno).all()
        asistentes_ids = {
            a.alumno_id for a in session.query(Asistencia).filter_by(sesion_id=active_session_id).all()
        }
        
        for alumno in todos_alumnos:
            if alumno.id not in asistentes_ids:
                falta = Asistencia(
                    sesion_id=active_session_id,
                    alumno_id=alumno.id,
                    metodo="AUTO",
                    estado="FALTA"
                )
                session.add(falta)
        session.commit()

        # 2. Consultar lista completa
        asistentes = (
            session.query(Asistencia, Alumno)
            .join(Alumno, Asistencia.alumno_id == Alumno.id)
            .filter(Asistencia.sesion_id == active_session_id)
            .order_by(Asistencia.timestamp)
            .all()
        )

        total_alumnos = len(asistentes)
        presentes = [a for a, _ in asistentes if a.estado == "PRESENTE"]
        tardes = [a for a, _ in asistentes if a.estado in ("TARDE", "RETARDO")]
        ausentes = [a for a, _ in asistentes if a.estado == "FALTA"]

        fecha_str = sesion_db.fecha.strftime("%d/%m/%Y") if sesion_db.fecha else datetime.date.today().strftime("%d/%m/%Y")
        h_apert = hora_apertura.strftime("%H:%M:%S") if hora_apertura else "—"
        h_cierr = datetime.datetime.now().strftime("%H:%M:%S")

        # 3. Construir reporte Telegram
        report = (
            f"📝 *REPORTE DE ASISTENCIA — REC*\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n"
            f"📖 *Materia:* {active_class_name}\n"
            f"📅 *Fecha:* {fecha_str}\n"
            f"⏰ *Apertura:* {h_apert} | *Cierre:* {h_cierr}\n"
            f"📊 *Sesión ID:* #{active_session_id}\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"📈 *Métricas:*\n"
            f"✅ Presentes: {len(presentes)}\n"
            f"⚠️ Retardos: {len(tardes)}\n"
            f"❌ Ausentes: {len(ausentes)}\n"
            f"👥 Total enrolados: {total_alumnos}\n\n"
            f"📋 *Detalle completo:*\n"
        )

        tabla = "```\n"
        tabla += "N°  | Código | Estudiante        | Hora     | Estado  \n"
        tabla += "----|--------|-------------------|----------|----------\n"
        for idx, (asis, alum) in enumerate(asistentes, start=1):
            nombre_limpio = alum.nombre.strip()
            nombre_fmt = (nombre_limpio[:14] + "...") if len(nombre_limpio) > 17 else nombre_limpio.ljust(17)
            t_str = asis.timestamp.strftime("%H:%M:%S") if (asis.timestamp and asis.estado != "FALTA") else "  —   "
            estado_fmt = asis.estado.ljust(8)
            tabla += f"{str(idx).zfill(2)}  | {str(asis.alumno_id).ljust(6)} | {nombre_fmt} | {t_str} | {estado_fmt}\n"
        tabla += "```\n"
        report += tabla
        report += f"━━━━━━━━━━━━━━━━━━━━━━\n🤖 _Sistema REC_"

        # 4. Guardar reporte .txt
        report_dir = BASE_DIR / "reportes"
        report_dir.mkdir(exist_ok=True)
        report_path = report_dir / f"reporte_sesion_{active_session_id}.txt"
        with open(report_path, "w", encoding="utf-8") as f:
            f.write(report.replace("*", "").replace("```", ""))

        # 5. Exportar a Excel
        if EXCEL_DISPONIBLE:
            export_excel_web(report_dir, asistentes, fecha_str, h_apert, h_cierr)

        # 6. Enviar a Telegram
        if TELEGRAM_BOT_TOKEN and TELEGRAM_CHAT_ID:
            def _send():
                try:
                    bot = telebot.TeleBot(TELEGRAM_BOT_TOKEN)
                    bot.send_message(TELEGRAM_CHAT_ID, report, parse_mode="Markdown")
                except Exception as tx:
                    print(f"[Telegram ERROR] {tx}")
            threading.Thread(target=_send, daemon=True).start()

    except Exception as e:
        print(f"Error al generar reporte: {e}")
    finally:
        session.close()


def export_excel_web(report_dir, asistentes, fecha_str, h_apert, h_cierr):
    """Genera la hoja de cálculo Excel formal con diseño elegante."""
    global active_class_name, active_session_id
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asistencia"
    ws.views.sheetView[0].showGridLines = True

    # Colores elegantes
    fill_header = PatternFill(start_color="1E1E35", end_color="1E1E35", fill_type="solid")
    fill_present = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    fill_late = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    fill_absent = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    font_title = Font(name="Segoe UI", size=16, bold=True, color="1E1E35")
    font_bold = Font(name="Segoe UI", size=11, bold=True)
    font_normal = Font(name="Segoe UI", size=11)
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    border_thin = Side(border_style="thin", color="D1D5DB")
    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    # Título y Metadata
    ws["B2"] = "REPORTE OFICIAL DE ASISTENCIA"
    ws["B2"].font = font_title
    ws.row_dimensions[2].height = 30

    metadata = [
        ("Materia:", active_class_name),
        ("Fecha:", fecha_str),
        ("Hora Apertura:", h_apert),
        ("Hora Cierre:", h_cierr),
        ("ID Sesión:", f"#{active_session_id}")
    ]

    for idx, (label, val) in enumerate(metadata, start=4):
        ws[f"B{idx}"] = label
        ws[f"B{idx}"].font = font_bold
        ws[f"C{idx}"] = val
        ws[f"C{idx}"].font = font_normal
        ws.row_dimensions[idx].height = 20

    # Cabeceras de Tabla
    headers = ["N°", "Código", "Estudiante", "Hora Registro", "Método", "Estado"]
    cols = ["B", "C", "D", "E", "F", "G"]
    
    ws.row_dimensions[10].height = 25
    for col_letter, header_text in zip(cols, headers):
        cell = ws[f"{col_letter}10"]
        cell.value = header_text
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = cell_border

    # Cargar Datos
    for row_idx, (asis, alum) in enumerate(asistentes, start=11):
        ws.row_dimensions[row_idx].height = 22
        
        ws[f"B{row_idx}"] = row_idx - 10
        ws[f"C{row_idx}"] = alum.id
        ws[f"D{row_idx}"] = alum.nombre
        ws[f"E{row_idx}"] = asis.timestamp.strftime("%H:%M:%S") if (asis.timestamp and asis.estado != "FALTA") else "—"
        ws[f"F{row_idx}"] = asis.metodo
        ws[f"G{row_idx}"] = asis.estado

        # Alinear celdas
        ws[f"B{row_idx}"].alignment = align_center
        ws[f"C{row_idx}"].alignment = align_center
        ws[f"D{row_idx}"].alignment = align_left
        ws[f"E{row_idx}"].alignment = align_center
        ws[f"F{row_idx}"].alignment = align_center
        ws[f"G{row_idx}"].alignment = align_center

        # Aplicar bordes
        for col_letter in cols:
            ws[f"{col_letter}{row_idx}"].border = cell_border
            ws[f"{col_letter}{row_idx}"].font = font_normal

        # Colorear según el Estado
        status_cell = ws[f"G{row_idx}"]
        if asis.estado == "PRESENTE":
            status_cell.fill = fill_present
        elif asis.estado in ("TARDE", "RETARDO"):
            status_cell.fill = fill_late
        else:
            status_cell.fill = fill_absent

    # Ajustar Ancho de Columnas
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 15

    excel_path = report_dir / f"reporte_sesion_{active_session_id}.xlsx"
    wb.save(excel_path)
    print(f"[Excel] Guardado en: {excel_path}")


if __name__ == "__main__":
    import uvicorn
    # Iniciar localmente en puerto 8000
    uvicorn.run(app, host="0.0.0.0", port=8000)
